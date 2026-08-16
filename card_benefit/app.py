# -*- coding: utf-8 -*-
"""카드 남은혜택 체크 - PC/폰 공용 웹앱.

같은 와이파이(공유기)에 연결된 PC와 폰이 브라우저로 같은 서버(0.0.0.0)에
접속해서 카드별 혜택의 이번 달 남은 한도를 같이 보고 같이 기록한다.
"""
import json
import os
import secrets
from datetime import date, datetime, timedelta
from io import BytesIO

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook

from models import (
    annotate_merchant,
    card_codes_match,
    compute_change_earned,
    compute_percent_discount,
    current_year_month,
    cycle_bounds,
    extract_last4,
    find_statement_header,
    get_conn,
    get_setting,
    init_db,
    is_cancelled_status,
    match_benefit_keyword,
    match_rate_table,
    parse_rate_table_categories,
    rate_table_entry_by_label,
    parse_amount_cell,
    parse_date_cell,
    parse_datetime_cell,
    parse_notification,
    previous_year_month,
    set_setting,
    tier_limit_for_spend,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

app = Flask(__name__)


def serialize_inbox(conn):
    rows = conn.execute(
        "SELECT * FROM inbox_items WHERE status = 'pending' ORDER BY occurred_at DESC, id DESC"
    ).fetchall()
    cards_with_code = conn.execute("SELECT id, name, last4 FROM cards WHERE last4 != ''").fetchall()
    benefits_by_card = {}
    for b in conn.execute(
        "SELECT id, card_id, name, merchant_keywords, calc_mode, always_doubled FROM benefits"
    ):
        benefits_by_card.setdefault(b["card_id"], []).append(b)

    items = []
    for r in rows:
        card = None
        if r["last4"]:
            for c in cards_with_code:
                if card_codes_match(c["last4"], r["last4"]):
                    card = c
                    break

        matched_benefit = None
        matched_doubled = False

        if card and r["merchant"]:
            for b in benefits_by_card.get(card["id"], []):
                if match_benefit_keyword(r["merchant"], b["merchant_keywords"]):
                    matched_benefit = b
                    matched_doubled = b["calc_mode"] == "change_under_1000"
                    break

        if not matched_benefit and card:
            # 가맹점 키워드로 못 찾았어도, 이 카드에 "잔돈 자동계산"형 혜택이
            # 있으면(더모아처럼 거의 모든 결제가 적립 대상인 카드) 기본으로
            # 그 혜택을 제안한다. 키워드가 비어있는(catch-all) 혜택이 있으면
            # 그걸 우선 쓰고("특별적립 전용" 혜택을 따로 등록해둔 경우 대비),
            # 없으면(대부분의 경우 - 혜택이 하나뿐인 경우) 아무 change_under_1000
            # 혜택이나 후보로 삼는다 - 혜택을 하나만 등록한 사람도 자동매칭돼야 하므로.
            change_benefits = [
                b for b in benefits_by_card.get(card["id"], []) if b["calc_mode"] == "change_under_1000"
            ]
            catch_all = [b for b in change_benefits if not b["merchant_keywords"]]
            fallback_pool = catch_all or change_benefits
            if fallback_pool:
                matched_benefit = fallback_pool[0]
                # merchant_keywords가 있는 혜택은 "매칭되면 2배, 안 되면 1배"를
                # 구분하려고 키워드를 등록해둔 것이므로, 여기(키워드 매칭 실패)로
                # 온 이상 always_doubled를 적용하면 안 된다. always_doubled는
                # 키워드가 아예 없는(=구분할 필요 없이 통째로 2배인) 혜택에만 쓴다.
                matched_doubled = bool(matched_benefit["always_doubled"]) and not matched_benefit["merchant_keywords"]

        # 카드사 엑셀의 "이용구분"에 해외/할부라고 명시돼 있으면, 가맹점명으로
        # 유추할 필요 없이 그 자체로 확실한 신호이므로 더모아형(잔돈 적립) 혜택은
        # 무조건 특별적립(2배) 대상으로 본다(안내장 기준: 해외이용, 할부거래는
        # 항상 2배).
        if matched_benefit and matched_benefit["calc_mode"] == "change_under_1000":
            usage_type = r["usage_type"] or ""
            if "해외" in usage_type or "할부" in usage_type:
                matched_doubled = True

        items.append({
            "id": r["id"],
            "raw_text": r["raw_text"],
            "amount": r["amount"],
            "last4": r["last4"],
            "issuer": r["issuer"],
            "merchant": r["merchant"],
            "occurred_at": r["occurred_at"],
            "matched_card_id": card["id"] if card else None,
            "matched_card_name": card["name"] if card else None,
            "matched_benefit_id": matched_benefit["id"] if matched_benefit else None,
            "matched_benefit_name": matched_benefit["name"] if matched_benefit else None,
            "matched_benefit_doubled": matched_doubled,
        })
    return items


def serialize_no_benefit(conn):
    rows = conn.execute(
        """SELECT inbox_items.*, cards.name AS card_name
           FROM inbox_items LEFT JOIN cards ON cards.id = inbox_items.card_id
           WHERE inbox_items.status = 'no_benefit'
           ORDER BY inbox_items.occurred_at DESC, inbox_items.id DESC"""
    ).fetchall()
    return [
        {
            "id": r["id"],
            "raw_text": r["raw_text"],
            "amount": r["amount"],
            "merchant": r["merchant"],
            "occurred_at": r["occurred_at"],
            "card_name": r["card_name"],
        }
        for r in rows
    ]


def serialize_full():
    conn = get_conn()
    try:
        return {
            "cards": _serialize_cards(conn),
            "inbox": serialize_inbox(conn),
            "no_benefit": serialize_no_benefit(conn),
        }
    finally:
        conn.close()


def _serialize_cards(conn):
    cards = conn.execute(
        "SELECT * FROM cards ORDER BY sort_order ASC, id ASC"
    ).fetchall()

    result = []
    for card in cards:
        start, end = cycle_bounds(card["reset_day"])

        this_month = current_year_month()
        prev_month = previous_year_month()
        perf_threshold = card["perf_threshold"] or 0

        # 실적은 손으로 입력할 필요 없이, 그동안 받은 결제 알림(배정됐거나
        # "혜택 없음" 처리된 것 - 즉 실제로 확인된 결제) 금액을 합산해서 자동
        # 계산한다. 카드사 실적은 "지난달(전월)" 완결된 금액을 기준으로 이번 달
        # 혜택 구간이 정해지므로, 지난달 알림 합계를 그 기준으로 쓴다.
        auto_prev_spend = conn.execute(
            """SELECT COALESCE(SUM(amount), 0) AS total FROM inbox_items
               WHERE card_id = ? AND status IN ('assigned', 'no_benefit')
               AND substr(occurred_at, 1, 7) = ?""",
            (card["id"], prev_month),
        ).fetchone()["total"]
        auto_this_month_spend = conn.execute(
            """SELECT COALESCE(SUM(amount), 0) AS total FROM inbox_items
               WHERE card_id = ? AND status IN ('assigned', 'no_benefit')
               AND substr(occurred_at, 1, 7) = ?""",
            (card["id"], this_month),
        ).fetchone()["total"]

        manual_row = conn.execute(
            "SELECT total_spend FROM performance WHERE card_id = ? AND year_month = ?",
            (card["id"], prev_month),
        ).fetchone()
        perf_spend = manual_row["total_spend"] if manual_row else auto_prev_spend

        benefits = conn.execute(
            "SELECT * FROM benefits WHERE card_id = ? ORDER BY sort_order ASC, id ASC",
            (card["id"],),
        ).fetchall()

        benefit_list = []
        for b in benefits:
            used = conn.execute(
                """SELECT COALESCE(SUM(used_value), 0) AS total FROM usage_logs
                   WHERE benefit_id = ? AND used_at >= ? AND used_at < ?""",
                (b["id"], start.isoformat(), end.isoformat()),
            ).fetchone()["total"]

            logs = conn.execute(
                """SELECT usage_logs.*, inbox_items.occurred_at AS notif_occurred_at
                   FROM usage_logs
                   LEFT JOIN inbox_items ON inbox_items.id = usage_logs.source_inbox_id
                   WHERE usage_logs.benefit_id = ?
                   AND usage_logs.used_at >= ? AND usage_logs.used_at < ?
                   ORDER BY usage_logs.used_at DESC, usage_logs.id DESC""",
                (b["id"], start.isoformat(), end.isoformat()),
            ).fetchall()

            tiered_limit = tier_limit_for_spend(b["tier_table"], perf_spend)
            limit_value = tiered_limit if tiered_limit is not None else (b["limit_value"] or 0)
            unlimited = limit_value <= 0
            remaining = None if unlimited else limit_value - used
            percent = None if unlimited else max(0, min(100, round(used / limit_value * 100))) if limit_value > 0 else 0

            category_usage = None
            if b["calc_mode"] == "percent_discount" and b["rate_table"]:
                categories = parse_rate_table_categories(b["rate_table"])
                if categories:
                    today_str = date.today().isoformat()
                    counts = {}
                    today_counts = {}
                    for l in logs:
                        cat = l["rate_category"]
                        if cat:
                            counts[cat] = counts.get(cat, 0) + 1
                            if l["used_at"] == today_str:
                                today_counts[cat] = today_counts.get(cat, 0) + 1
                    category_usage = [
                        {
                            "label": c["label"],
                            "count": counts.get(c["label"], 0),
                            "today_count": today_counts.get(c["label"], 0),
                            "daily_limit": c["daily_limit"],
                            "monthly_limit": c["monthly_limit"],
                        }
                        for c in categories
                    ]

            benefit_list.append({
                "id": b["id"],
                "card_id": b["card_id"],
                "name": b["name"],
                "limit_type": b["limit_type"],
                "limit_value": limit_value,
                "unlimited": unlimited,
                "tiered": tiered_limit is not None,
                "tier_table": b["tier_table"],
                "memo": b["memo"],
                "merchant_keywords": b["merchant_keywords"],
                "calc_mode": b["calc_mode"],
                "discount_percent": b["discount_percent"],
                "per_txn_cap": b["per_txn_cap"],
                "rate_table": b["rate_table"],
                "category_usage": category_usage,
                "always_doubled": bool(b["always_doubled"]),
                "used": used,
                "remaining": remaining,
                "percent": percent,
                "over_limit": (not unlimited) and remaining < 0,
                "logs": [dict(l) for l in logs],
            })

        days_left = (end - date.today()).days

        result.append({
            "id": card["id"],
            "name": card["name"],
            "issuer": card["issuer"],
            "last4": card["last4"],
            "reset_day": card["reset_day"],
            "perf_threshold": perf_threshold,
            "perf_spend": perf_spend,
            "perf_month": prev_month,
            "perf_auto": manual_row is None,
            "perf_met": (auto_this_month_spend >= perf_threshold) if perf_threshold > 0 else None,
            "this_month_spend": auto_this_month_spend,
            "memo": card["memo"],
            "cycle_start": start.isoformat(),
            "cycle_end": (end - timedelta(days=1)).isoformat(),
            "days_left": days_left,
            "benefits": benefit_list,
        })
    return result


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/state")
def api_state():
    return jsonify(serialize_full())


@app.route("/api/cards", methods=["POST"])
def create_card():
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "카드 이름을 입력하세요."}), 400
    try:
        reset_day = int(data.get("reset_day") or 1)
    except (TypeError, ValueError):
        reset_day = 1
    try:
        perf_threshold = float(data.get("perf_threshold") or 0)
    except (TypeError, ValueError):
        perf_threshold = 0

    conn = get_conn()
    try:
        max_order = conn.execute("SELECT COALESCE(MAX(sort_order), -1) AS m FROM cards").fetchone()["m"]
        conn.execute(
            """INSERT INTO cards (name, issuer, last4, reset_day, perf_threshold, memo, sort_order)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                name,
                (data.get("issuer") or "").strip(),
                (data.get("last4") or "").strip(),
                reset_day,
                perf_threshold,
                (data.get("memo") or "").strip(),
                max_order + 1,
            ),
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


@app.route("/api/cards/<int:card_id>", methods=["PUT"])
def update_card(card_id):
    data = request.get_json(force=True) or {}
    conn = get_conn()
    try:
        card = conn.execute("SELECT * FROM cards WHERE id = ?", (card_id,)).fetchone()
        if not card:
            return jsonify({"error": "카드를 찾을 수 없습니다."}), 404

        name = (data.get("name") or card["name"]).strip()
        try:
            reset_day = int(data.get("reset_day", card["reset_day"]))
        except (TypeError, ValueError):
            reset_day = card["reset_day"]
        try:
            perf_threshold = float(data.get("perf_threshold", card["perf_threshold"]))
        except (TypeError, ValueError):
            perf_threshold = card["perf_threshold"]

        conn.execute(
            """UPDATE cards SET name = ?, issuer = ?, last4 = ?, reset_day = ?, perf_threshold = ?, memo = ?
               WHERE id = ?""",
            (
                name,
                (data.get("issuer", card["issuer"]) or "").strip(),
                (data.get("last4", card["last4"]) or "").strip(),
                reset_day,
                perf_threshold,
                (data.get("memo", card["memo"]) or "").strip(),
                card_id,
            ),
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


@app.route("/api/cards/<int:card_id>/performance", methods=["POST"])
def update_performance(card_id):
    """전월실적 계산용 "이번 달 총 사용액"을 직접 입력/수정한다.

    혜택별 사용기록과 달리, 전월실적은 보통 그 카드로 쓴 전체 금액을 보므로
    (스타벅스 할인 같은 특정 혜택에 해당 안 하는 결제도 포함) 카드사 앱/문자로
    확인한 "이번 달 실적" 숫자를 그대로 입력하는 방식으로 관리한다.
    """
    data = request.get_json(force=True) or {}
    try:
        total_spend = float(data.get("total_spend"))
    except (TypeError, ValueError):
        return jsonify({"error": "사용액을 입력하세요."}), 400
    if total_spend < 0:
        return jsonify({"error": "사용액은 0 이상이어야 합니다."}), 400

    year_month = (data.get("year_month") or "").strip() or previous_year_month()

    conn = get_conn()
    try:
        card = conn.execute("SELECT id FROM cards WHERE id = ?", (card_id,)).fetchone()
        if not card:
            return jsonify({"error": "카드를 찾을 수 없습니다."}), 404
        conn.execute(
            """INSERT INTO performance (card_id, year_month, total_spend, updated_at)
               VALUES (?, ?, ?, datetime('now', 'localtime'))
               ON CONFLICT(card_id, year_month)
               DO UPDATE SET total_spend = excluded.total_spend, updated_at = excluded.updated_at""",
            (card_id, year_month, total_spend),
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


@app.route("/api/cards/<int:card_id>", methods=["DELETE"])
def delete_card(card_id):
    conn = get_conn()
    try:
        conn.execute("DELETE FROM cards WHERE id = ?", (card_id,))
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


def _clean_tier_table(raw):
    """[[기준금액,한도], ...] 형태인지 검증하고 정돈된 JSON 문자열(또는 '')을 돌려준다."""
    raw = (raw or "").strip()
    if not raw:
        return ""
    try:
        tiers = json.loads(raw)
        if not isinstance(tiers, list):
            raise ValueError
        cleaned = [[float(t[0]), float(t[1])] for t in tiers]
    except (ValueError, TypeError, IndexError, KeyError):
        raise ValueError("구간표 형식이 올바르지 않습니다. 예: [[400000,5000],[800000,10000]]")
    return json.dumps(cleaned)


def _clean_rate_table(raw):
    """[["키워드1,키워드2",할인율], ...] 또는 [["키워드",할인율,건당한도], ...] 또는
    [["키워드",할인율,건당한도,일한도,월한도], ...] 형태인지 검증하고 정돈된 JSON
    문자열(또는 '')을 돌려준다."""
    raw = (raw or "").strip()
    if not raw:
        return ""
    try:
        rows = json.loads(raw)
        if not isinstance(rows, list):
            raise ValueError
        cleaned = []
        for r in rows:
            if len(r) not in (2, 3, 4, 5):
                raise ValueError
            row = [str(r[0]), float(r[1])]
            for extra in r[2:]:
                row.append(float(extra))
            cleaned.append(row)
    except (ValueError, TypeError, IndexError, KeyError):
        raise ValueError(
            "업종별 할인율 형식이 올바르지 않습니다. "
            '예: [["쇼핑몰,백화점,이커머스",10],["버스,지하철,택시",5]] '
            '(건당한도는 3번째 값, 일/월 횟수한도는 4·5번째 값으로 추가 가능: '
            '[["쇼핑",10,50000,1,5]])'
        )
    return json.dumps(cleaned, ensure_ascii=False)


@app.route("/api/cards/<int:card_id>/benefits", methods=["POST"])
def create_benefit(card_id):
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "혜택 이름을 입력하세요."}), 400
    try:
        limit_value = float(data.get("limit_value") or 0)
    except (TypeError, ValueError):
        limit_value = 0
    limit_type = data.get("limit_type") if data.get("limit_type") in ("amount", "count") else "amount"
    calc_mode = data.get("calc_mode") if data.get("calc_mode") in ("raw", "change_under_1000", "percent_discount") else "raw"
    try:
        discount_percent = float(data.get("discount_percent") or 0)
    except (TypeError, ValueError):
        discount_percent = 0
    try:
        per_txn_cap = float(data.get("per_txn_cap") or 0)
    except (TypeError, ValueError):
        per_txn_cap = 0
    try:
        tier_table = _clean_tier_table(data.get("tier_table"))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    try:
        rate_table = _clean_rate_table(data.get("rate_table"))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    always_doubled = 1 if data.get("always_doubled") else 0

    conn = get_conn()
    try:
        card = conn.execute("SELECT id FROM cards WHERE id = ?", (card_id,)).fetchone()
        if not card:
            return jsonify({"error": "카드를 찾을 수 없습니다."}), 404
        max_order = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) AS m FROM benefits WHERE card_id = ?", (card_id,)
        ).fetchone()["m"]
        conn.execute(
            """INSERT INTO benefits (card_id, name, limit_type, limit_value, memo, merchant_keywords, tier_table,
               calc_mode, discount_percent, per_txn_cap, rate_table, always_doubled, sort_order)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                card_id, name, limit_type, limit_value,
                (data.get("memo") or "").strip(),
                (data.get("merchant_keywords") or "").strip(),
                tier_table,
                calc_mode,
                discount_percent,
                per_txn_cap,
                rate_table,
                always_doubled,
                max_order + 1,
            ),
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


@app.route("/api/benefits/<int:benefit_id>", methods=["PUT"])
def update_benefit(benefit_id):
    data = request.get_json(force=True) or {}
    conn = get_conn()
    try:
        b = conn.execute("SELECT * FROM benefits WHERE id = ?", (benefit_id,)).fetchone()
        if not b:
            return jsonify({"error": "혜택을 찾을 수 없습니다."}), 404

        name = (data.get("name") or b["name"]).strip()
        try:
            limit_value = float(data.get("limit_value", b["limit_value"]))
        except (TypeError, ValueError):
            limit_value = b["limit_value"]
        limit_type = data.get("limit_type") if data.get("limit_type") in ("amount", "count") else b["limit_type"]
        calc_mode = data.get("calc_mode") if data.get("calc_mode") in ("raw", "change_under_1000", "percent_discount") else b["calc_mode"]
        try:
            discount_percent = float(data.get("discount_percent", b["discount_percent"]))
        except (TypeError, ValueError):
            discount_percent = b["discount_percent"]
        try:
            per_txn_cap = float(data.get("per_txn_cap", b["per_txn_cap"]))
        except (TypeError, ValueError):
            per_txn_cap = b["per_txn_cap"]
        try:
            tier_table = _clean_tier_table(data.get("tier_table", b["tier_table"]))
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        try:
            rate_table = _clean_rate_table(data.get("rate_table", b["rate_table"]))
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        always_doubled = 1 if data.get("always_doubled", b["always_doubled"]) else 0

        conn.execute(
            """UPDATE benefits SET name = ?, limit_type = ?, limit_value = ?, memo = ?, merchant_keywords = ?,
               tier_table = ?, calc_mode = ?, discount_percent = ?, per_txn_cap = ?, rate_table = ?,
               always_doubled = ? WHERE id = ?""",
            (
                name, limit_type, limit_value,
                (data.get("memo", b["memo"]) or "").strip(),
                (data.get("merchant_keywords", b["merchant_keywords"]) or "").strip(),
                tier_table,
                calc_mode,
                discount_percent,
                per_txn_cap,
                rate_table,
                always_doubled,
                benefit_id,
            ),
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


@app.route("/api/benefits/<int:benefit_id>", methods=["DELETE"])
def delete_benefit(benefit_id):
    conn = get_conn()
    try:
        conn.execute("DELETE FROM benefits WHERE id = ?", (benefit_id,))
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


def _duplicate_merchant_today(conn, benefit_id, used_at, merchant):
    """더모아형 혜택은 "동일 가맹점 1일 1회"만 적립되므로, 같은 날 같은
    가맹점으로 이미 기록이 있으면 그 뒤 결제는 적립 대상에서 제외한다."""
    if not merchant:
        return False
    row = conn.execute(
        """SELECT 1 FROM usage_logs WHERE benefit_id = ? AND used_at = ? AND merchant = ?
           LIMIT 1""",
        (benefit_id, used_at, merchant),
    ).fetchone()
    return row is not None


def _category_usage_count(conn, benefit_id, category, used_at, monthly=False):
    """rate_table 업종(category)별로 오늘(또는 이번 달) 몇 번 사용됐는지 센다.
    Daily Plan처럼 "업종별 일 1회, 월 5회까지"인 경우 한도 초과 여부 확인에 쓴다."""
    if monthly:
        row = conn.execute(
            """SELECT COUNT(*) AS c FROM usage_logs
               WHERE benefit_id = ? AND rate_category = ? AND substr(used_at, 1, 7) = ?""",
            (benefit_id, category, used_at[:7]),
        ).fetchone()
    else:
        row = conn.execute(
            "SELECT COUNT(*) AS c FROM usage_logs WHERE benefit_id = ? AND rate_category = ? AND used_at = ?",
            (benefit_id, category, used_at),
        ).fetchone()
    return row["c"]


def _apply_calc_mode(conn, benefit_id, benefit, raw_value, doubled, memo, merchant="", used_at="", category_override=""):
    """혜택의 calc_mode에 따라 입력값(결제금액 등)을 실제 한도 소진값으로 변환한다.

    "raw"는 입력값을 그대로 쓰고, 나머지는 결제금액을 자동으로 실제
    혜택금액(잔돈/할인액)으로 환산한다 - 결제금액 총액이 그대로 한도를
    깎아먹지 않도록. (used_value, memo, rate_category, error) 를 돌려준다.
    error가 있으면(업종별 일/월 횟수 한도 초과 등) 기록하지 않고 그 문구를 보여준다.
    category_override가 있으면(네이버페이/토스페이처럼 가맹점명으로 업종을 알 수
    없어서 사용자가 직접 고른 경우) 가맹점명 매칭 대신 그 업종 값을 그대로 쓴다.
    """
    if benefit["calc_mode"] == "change_under_1000":
        doubled = doubled or bool(benefit["always_doubled"])
        earned = compute_change_earned(raw_value, doubled)
        parts = [f"결제 {raw_value:,.0f}원 → 잔돈 {earned:,.0f}원{'(2배)' if doubled else ''}"]
        if memo:
            parts.append(memo)
        return earned, " · ".join(parts), "", None

    if benefit["calc_mode"] == "percent_discount":
        if category_override:
            percent, cap, daily_limit, monthly_limit = rate_table_entry_by_label(
                benefit["rate_table"], category_override, benefit["discount_percent"], benefit["per_txn_cap"]
            )
            category = category_override
        else:
            percent, cap, daily_limit, monthly_limit, category = match_rate_table(
                merchant, benefit["rate_table"], benefit["discount_percent"], benefit["per_txn_cap"]
            )
        if category and used_at:
            if daily_limit and _category_usage_count(conn, benefit_id, category, used_at) >= daily_limit:
                return None, None, None, f"이 업종은 하루 {daily_limit:g}회까지만 할인되는데, 오늘 이미 한도를 채워서 이 결제는 할인 대상이 아닙니다."
            if monthly_limit and _category_usage_count(conn, benefit_id, category, used_at, monthly=True) >= monthly_limit:
                return None, None, None, f"이 업종은 이번 달 {monthly_limit:g}회까지만 할인되는데, 이미 한도를 채워서 이 결제는 할인 대상이 아닙니다."
        earned = compute_percent_discount(raw_value, percent, cap)
        category_label = f"{category} " if category else ""
        parts = [f"결제 {raw_value:,.0f}원 → 할인 {earned:,.0f}원({category_label}{percent:g}%)"]
        if memo:
            parts.append(memo)
        return earned, " · ".join(parts), (category or ""), None

    return raw_value, memo, "", None


@app.route("/api/benefits/<int:benefit_id>/use", methods=["POST"])
def log_usage(benefit_id):
    data = request.get_json(force=True) or {}
    conn = get_conn()
    try:
        b = conn.execute("SELECT * FROM benefits WHERE id = ?", (benefit_id,)).fetchone()
        if not b:
            return jsonify({"error": "혜택을 찾을 수 없습니다."}), 404

        try:
            raw_value = float(data.get("used_value"))
        except (TypeError, ValueError):
            raw_value = 1 if b["limit_type"] == "count" else 0
        if raw_value == 0:
            return jsonify({"error": "사용 금액(또는 횟수)을 입력하세요."}), 400

        used_at = (data.get("used_at") or "").strip()
        try:
            datetime.strptime(used_at, "%Y-%m-%d")
        except ValueError:
            used_at = date.today().isoformat()

        merchant = (data.get("merchant") or "").strip()
        memo = (data.get("memo") or "").strip()
        category_override = (data.get("category") or "").strip()

        if b["calc_mode"] == "change_under_1000" and _duplicate_merchant_today(conn, benefit_id, used_at, merchant):
            return jsonify({
                "error": f"같은 날 \"{merchant}\"에 이미 기록이 있습니다. "
                         "동일 가맹점은 1일 1회만 적립되므로(가장 먼저 결제한 건만 인정), "
                         "이 결제는 적립 대상이 아닙니다."
            }), 400

        used_value, memo, rate_category, error = _apply_calc_mode(
            conn, benefit_id, b, raw_value, bool(data.get("doubled")), memo, merchant, used_at, category_override
        )
        if error:
            return jsonify({"error": error}), 400

        conn.execute(
            """INSERT INTO usage_logs (benefit_id, used_value, used_at, merchant, memo, rate_category)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (benefit_id, used_value, used_at, merchant, memo, rate_category),
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


@app.route("/api/logs/<int:log_id>", methods=["DELETE"])
def delete_log(log_id):
    conn = get_conn()
    try:
        log = conn.execute("SELECT source_inbox_id FROM usage_logs WHERE id = ?", (log_id,)).fetchone()
        conn.execute("DELETE FROM usage_logs WHERE id = ?", (log_id,))
        if log and log["source_inbox_id"]:
            # 받은 알림에서 배정되어 만들어진 기록이면, 삭제할 때 알림을
            # 없애버리지 않고 "받은 결제 알림" 목록으로 되돌린다(다시 배정 가능).
            conn.execute(
                "UPDATE inbox_items SET status = 'pending', card_id = NULL, benefit_id = NULL WHERE id = ?",
                (log["source_inbox_id"],),
            )
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


# ---- 폰 결제 알림 자동수집 (MacroDroid 등에서 웹훅으로 호출) ----


def _get_or_create_inbox_token(conn):
    token = get_setting(conn, "inbox_token")
    if not token:
        token = secrets.token_hex(4)
        set_setting(conn, "inbox_token", token)
        conn.commit()
    return token


def _inbox_authorized(req, expected_token):
    supplied = req.args.get("token") or req.headers.get("X-Inbox-Token")
    if not supplied:
        supplied = (req.get_json(silent=True) or {}).get("token")
    return supplied == expected_token


@app.route("/api/settings/inbox", methods=["GET"])
def get_inbox_settings():
    conn = get_conn()
    try:
        token = _get_or_create_inbox_token(conn)
    finally:
        conn.close()
    return jsonify({"token": token, "webhook_path": "/api/inbox"})


@app.route("/api/settings/inbox/regenerate", methods=["POST"])
def regenerate_inbox_token():
    conn = get_conn()
    try:
        token = secrets.token_hex(4)
        set_setting(conn, "inbox_token", token)
        conn.commit()
    finally:
        conn.close()
    return jsonify({"token": token, "webhook_path": "/api/inbox"})


@app.route("/api/inbox", methods=["POST"])
def create_inbox_item():
    data = request.get_json(force=True, silent=True) or {}
    text = (data.get("text") or "").strip()
    if not text:
        # 유효한 JSON이 아니거나 "text" 필드가 없으면, 본문 전체를 알림 원문으로
        # 취급한다. MacroDroid 등에서 알림 텍스트를 JSON으로 정확히 감싸서
        # 보내기 번거로운 경우, 그냥 알림 텍스트를 body에 그대로 넣어도 되게
        # 하기 위함(줄바꿈/따옴표 이스케이프를 신경 쓸 필요가 없어짐).
        text = (request.get_data(as_text=True) or "").strip()

    conn = get_conn()
    try:
        expected_token = _get_or_create_inbox_token(conn)
        if not _inbox_authorized(request, expected_token):
            return jsonify({"error": "인증 토큰이 올바르지 않습니다."}), 401

        if not text:
            return jsonify({"error": "text가 비어 있습니다."}), 400

        parsed = parse_notification(text)
        occurred_at = (data.get("occurred_at") or "").strip() or datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # MacroDroid 등에서 알림 하나에 트리거가 두 번 발동하는 경우가 있어서,
        # 같은 문구가 60초 안에 다시 들어오면 중복으로 보고 건너뛴다.
        recent_dup = conn.execute(
            """SELECT 1 FROM inbox_items WHERE raw_text = ?
               AND created_at >= datetime('now', 'localtime', '-60 seconds') LIMIT 1""",
            (text,),
        ).fetchone()
        if not recent_dup:
            conn.execute(
                """INSERT INTO inbox_items (raw_text, amount, last4, issuer, merchant, occurred_at)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (text, parsed["amount"], parsed["last4"], parsed["issuer"], parsed["merchant"], occurred_at),
            )
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


@app.route("/api/inbox/<int:item_id>/assign", methods=["POST"])
def assign_inbox_item(item_id):
    data = request.get_json(force=True) or {}
    try:
        benefit_id = int(data.get("benefit_id"))
    except (TypeError, ValueError):
        return jsonify({"error": "혜택을 선택하세요."}), 400

    conn = get_conn()
    try:
        item = conn.execute("SELECT * FROM inbox_items WHERE id = ?", (item_id,)).fetchone()
        if not item or item["status"] != "pending":
            return jsonify({"error": "이미 처리되었거나 존재하지 않는 알림입니다."}), 404

        benefit = conn.execute("SELECT * FROM benefits WHERE id = ?", (benefit_id,)).fetchone()
        if not benefit:
            return jsonify({"error": "혜택을 찾을 수 없습니다."}), 404

        try:
            raw_value = float(data.get("amount") if data.get("amount") not in (None, "") else item["amount"])
        except (TypeError, ValueError):
            raw_value = 0
        if not raw_value:
            return jsonify({"error": "금액을 확인할 수 없습니다. 직접 입력해주세요."}), 400

        used_at = (data.get("used_at") or "").strip() or (item["occurred_at"] or "")[:10]
        try:
            datetime.strptime(used_at, "%Y-%m-%d")
        except ValueError:
            used_at = date.today().isoformat()

        merchant = (data.get("merchant") or item["merchant"] or "").strip()
        memo = (data.get("memo") or "").strip()
        category_override = (data.get("category") or "").strip()

        if benefit["calc_mode"] == "change_under_1000" and _duplicate_merchant_today(conn, benefit_id, used_at, merchant):
            return jsonify({
                "error": f"같은 날 \"{merchant}\"에 이미 기록이 있습니다. "
                         "동일 가맹점은 1일 1회만 적립되므로(가장 먼저 결제한 건만 인정), "
                         "이 결제는 적립 대상이 아닙니다. 알림을 무시 처리해주세요."
            }), 400

        used_value, memo, rate_category, error = _apply_calc_mode(
            conn, benefit_id, benefit, raw_value, bool(data.get("doubled")), memo, merchant, used_at, category_override
        )
        if error:
            return jsonify({"error": error}), 400

        conn.execute(
            """INSERT INTO usage_logs (benefit_id, used_value, used_at, merchant, memo, source_inbox_id, rate_category)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (benefit_id, used_value, used_at, merchant, memo, item_id, rate_category),
        )
        conn.execute(
            "UPDATE inbox_items SET status = 'assigned', card_id = ?, benefit_id = ? WHERE id = ?",
            (benefit["card_id"], benefit_id, item_id),
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


@app.route("/api/inbox/<int:item_id>/mark-no-benefit", methods=["POST"])
def mark_no_benefit(item_id):
    """해당하는 혜택이 없는(그러나 실제로 결제는 한) 알림을 "받은 결제 알림"
    목록에서 빼되, 완전히 삭제(=이 알림 무시)하지 않고 별도 목록에 남겨둔다."""
    data = request.get_json(force=True, silent=True) or {}
    card_id = data.get("card_id") or None

    conn = get_conn()
    try:
        item = conn.execute("SELECT * FROM inbox_items WHERE id = ?", (item_id,)).fetchone()
        if not item or item["status"] != "pending":
            return jsonify({"error": "이미 처리되었거나 존재하지 않는 알림입니다."}), 404

        conn.execute(
            "UPDATE inbox_items SET status = 'no_benefit', card_id = ?, benefit_id = NULL WHERE id = ?",
            (card_id, item_id),
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


@app.route("/api/inbox/<int:item_id>/reopen", methods=["POST"])
def reopen_inbox_item(item_id):
    """"혜택 없음"으로 처리했던 알림을 다시 "받은 결제 알림" 목록으로 되돌린다."""
    conn = get_conn()
    try:
        conn.execute(
            "UPDATE inbox_items SET status = 'pending', card_id = NULL, benefit_id = NULL WHERE id = ?",
            (item_id,),
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


@app.route("/api/inbox/<int:item_id>", methods=["DELETE"])
def discard_inbox_item(item_id):
    conn = get_conn()
    try:
        conn.execute("DELETE FROM inbox_items WHERE id = ?", (item_id,))
        conn.commit()
    finally:
        conn.close()
    return jsonify(serialize_full())


# ---- 지난/누락 내역 백업(엑셀 다운로드) · 일괄 업로드 ----


def _xlsx_response(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename, mimetype=XLSX_MIME)


def _set_column_widths(ws, widths):
    for col, width in zip("ABCDEFGH", widths):
        ws.column_dimensions[col].width = width


@app.route("/api/export/usage.xlsx")
def export_usage():
    conn = get_conn()
    try:
        rows = conn.execute(
            """SELECT c.name AS card_name, b.name AS benefit_name, l.used_at, l.used_value, l.memo
               FROM usage_logs l
               JOIN benefits b ON b.id = l.benefit_id
               JOIN cards c ON c.id = b.card_id
               ORDER BY l.used_at DESC, l.id DESC"""
        ).fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "사용내역"
    ws.append(["카드명", "혜택명", "날짜", "금액(또는 횟수)", "메모"])
    for r in rows:
        ws.append([r["card_name"], r["benefit_name"], r["used_at"], r["used_value"], r["memo"]])
    _set_column_widths(ws, [18, 22, 12, 16, 26])

    return _xlsx_response(wb, "카드혜택_사용내역.xlsx")


@app.route("/api/export/template.xlsx")
def export_template():
    conn = get_conn()
    try:
        combos = conn.execute(
            """SELECT c.name AS card_name, b.name AS benefit_name
               FROM benefits b JOIN cards c ON c.id = b.card_id
               ORDER BY c.sort_order, b.sort_order"""
        ).fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "업로드양식"
    ws.append(["카드명", "혜택명", "날짜(YYYY-MM-DD)", "금액(또는 횟수)", "메모(선택)"])
    for c in combos:
        ws.append([c["card_name"], c["benefit_name"], "", "", ""])
    _set_column_widths(ws, [18, 22, 18, 16, 26])

    return _xlsx_response(wb, "카드혜택_업로드양식.xlsx")


def _import_own_template(ws):
    """이 앱이 만든 "빈 양식"(카드명/혜택명 헤더)을 그대로 채워 올린 경우.

    카드명·혜택명이 정확히 일치하는 행은 바로 usage_logs로 들어간다(혜택이
    이미 정해져 있으므로 사람이 다시 고를 필요가 없음).
    """
    conn = get_conn()
    try:
        benefit_lookup = {}
        for row in conn.execute(
            """SELECT b.id AS benefit_id, c.name AS card_name, b.name AS benefit_name
               FROM benefits b JOIN cards c ON c.id = b.card_id"""
        ):
            benefit_lookup[(row["card_name"].strip(), row["benefit_name"].strip())] = row["benefit_id"]

        # 같은 양식을 다시 올려도 중복으로 안 쌓이도록, 이미 있는 (혜택, 날짜,
        # 금액, 메모) 조합을 미리 읽어둔다.
        seen_combo = {
            (r["benefit_id"], r["used_at"], r["used_value"], r["memo"] or "")
            for r in conn.execute("SELECT benefit_id, used_at, used_value, memo FROM usage_logs")
        }

        added = 0
        skipped = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(v in (None, "") for v in row):
                continue
            card_name, benefit_name, used_at, used_value = (list(row) + [None] * 4)[:4]
            memo = row[4] if len(row) > 4 else None
            memo_str = str(memo).strip() if memo else ""

            key = ((str(card_name).strip() if card_name else ""), (str(benefit_name).strip() if benefit_name else ""))
            if key not in benefit_lookup:
                skipped.append(f"{i}행: 카드/혜택 이름을 찾을 수 없음 ({key[0]} / {key[1]})")
                continue

            used_value = parse_amount_cell(used_value)
            if not used_value:
                skipped.append(f"{i}행: 금액(또는 횟수)이 올바르지 않음")
                continue

            used_at_str = parse_date_cell(used_at)
            if not used_at_str:
                skipped.append(f"{i}행: 날짜가 올바르지 않음 (YYYY-MM-DD)")
                continue

            benefit_id = benefit_lookup[key]
            combo = (benefit_id, used_at_str, used_value, memo_str)
            if combo in seen_combo:
                skipped.append(f"{i}행: 같은 혜택·날짜·금액 내역이 이미 있어 중복으로 보고 건너뜀")
                continue
            seen_combo.add(combo)

            conn.execute(
                "INSERT INTO usage_logs (benefit_id, used_value, used_at, memo) VALUES (?, ?, ?, ?)",
                (benefit_id, used_value, used_at_str, memo_str),
            )
            added += 1
        conn.commit()
    finally:
        conn.close()

    result = serialize_full()
    result["import_result"] = {"mode": "template", "added": added, "skipped": skipped}
    return jsonify(result)


def _import_statement(ws, header_info):
    """카드사에서 그대로 다운로드한 엑셀(이용일자/승인금액 등)을 인식해서
    "받은 결제 알림"(inbox_items) 목록에 대량으로 쌓는다.

    이 표에는 어떤 혜택인지에 대한 정보가 없으므로, 자동수집 알림과 똑같이
    사람이 앱에서 "혜택 선택"으로 하나씩 배정해야 한다.
    """
    date_col = header_info["date_col"]
    amount_col = header_info["amount_col"]
    merchant_col = header_info["merchant_col"]
    cardno_col = header_info["cardno_col"]
    approval_col = header_info.get("approval_col")
    usage_type_col = header_info.get("usage_type_col")

    conn = get_conn()
    try:
        cards_with_code = conn.execute(
            """SELECT c.id, c.name, c.last4, COUNT(b.id) AS benefit_count
               FROM cards c LEFT JOIN benefits b ON b.card_id = c.id
               WHERE c.last4 != '' GROUP BY c.id"""
        ).fetchall()

        # 같은 엑셀을 실수로(또는 빠진 내역을 채우려고) 다시 올려도, 그리고
        # 이미 폰 알림(자동수집)으로 들어와 있는 같은 결제를 엑셀로 또 올려도
        # 중복으로 안 쌓이도록 막는다. 승인번호가 있으면 그걸로, 없으면
        # (카드뒤4자리+금액+날짜) 조합으로 막는다 - 가맹점명은 일부러 뺐다.
        # 폰 알림은 "스타벅스 강남점"처럼, 엑셀은 "주식회사 스타벅스코리아"처럼
        # 같은 결제를 서로 다른 문구로 남기는 경우가 많아서, 가맹점명까지
        # 똑같아야 막는 조건이면 정작 이런 교차 중복을 못 걸러낸다.
        seen_approval = {
            r["approval_no"] for r in conn.execute(
                "SELECT approval_no FROM inbox_items WHERE approval_no != ''"
            )
        }
        seen_combo = {
            (r["last4"], r["amount"], (r["occurred_at"] or "")[:10])
            for r in conn.execute("SELECT last4, amount, occurred_at FROM inbox_items")
        }

        queued = 0
        skipped = []
        for i, row in enumerate(
            ws.iter_rows(min_row=header_info["header_row"] + 1, values_only=True),
            start=header_info["header_row"] + 1,
        ):
            if row is None or all(v in (None, "") for v in row):
                continue

            status_col = header_info.get("status_col")
            if status_col is not None and status_col < len(row) and is_cancelled_status(row[status_col]):
                skipped.append(f"{i}행: 취소/실패 거래로 보여 건너뜀")
                continue

            amount = parse_amount_cell(row[amount_col] if amount_col < len(row) else None)
            if amount is None or amount <= 0:
                skipped.append(f"{i}행: 취소·환불로 보이는 금액이라 건너뜀" if amount is not None and amount < 0
                                else f"{i}행: 금액을 확인할 수 없음")
                continue

            used_at = parse_date_cell(row[date_col] if date_col < len(row) else None)
            if not used_at:
                skipped.append(f"{i}행: 날짜를 확인할 수 없음")
                continue
            occurred_at = parse_datetime_cell(row[date_col] if date_col < len(row) else None) or used_at

            merchant = ""
            if merchant_col is not None and merchant_col < len(row) and row[merchant_col] is not None:
                merchant = annotate_merchant(str(row[merchant_col]).strip())

            last4 = ""
            if cardno_col is not None and cardno_col < len(row):
                last4 = extract_last4(row[cardno_col])

            if last4:
                matched = next((c for c in cards_with_code if card_codes_match(c["last4"], last4)), None)
                if matched and matched["benefit_count"] == 0:
                    skipped.append(f"{i}행: \"{matched['name']}\"는 추적할 혜택이 등록되어 있지 않아 건너뜀")
                    continue

            approval_no = ""
            if approval_col is not None and approval_col < len(row) and row[approval_col] not in (None, ""):
                approval_no = str(row[approval_col]).strip()

            usage_type = ""
            if usage_type_col is not None and usage_type_col < len(row) and row[usage_type_col] not in (None, ""):
                usage_type = str(row[usage_type_col]).strip()

            if approval_no:
                if approval_no in seen_approval:
                    skipped.append(f"{i}행: 승인번호({approval_no})가 이미 등록되어 있어 중복으로 보고 건너뜀")
                    continue
                seen_approval.add(approval_no)
            else:
                combo = (last4, amount, used_at)
                if combo in seen_combo:
                    skipped.append(f"{i}행: 같은 카드·금액·날짜 내역이 이미 있어 중복으로 보고 건너뜀")
                    continue
                seen_combo.add(combo)

            raw_parts = [p for p in [merchant, f"{amount:,.0f}원", used_at] if p]
            raw_text = " · ".join(raw_parts)

            conn.execute(
                """INSERT INTO inbox_items (raw_text, amount, last4, issuer, merchant, occurred_at, approval_no, usage_type)
                   VALUES (?, ?, ?, '', ?, ?, ?, ?)""",
                (raw_text, amount, last4, merchant, occurred_at, approval_no, usage_type),
            )
            queued += 1
        conn.commit()
    finally:
        conn.close()

    result = serialize_full()
    result["import_result"] = {"mode": "statement", "queued": queued, "skipped": skipped}
    return jsonify(result)


@app.route("/api/import/usage", methods=["POST"])
def import_usage():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "파일을 선택하세요."}), 400

    try:
        wb = load_workbook(file, data_only=True)
    except Exception:
        return jsonify({"error": "엑셀 파일을 읽을 수 없습니다. .xlsx 파일인지 확인해주세요."}), 400
    ws = wb.active

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
    header_cells = [str(c).strip() if c is not None else "" for c in first_row]
    is_own_template = any("카드명" in c for c in header_cells) and any("혜택명" in c for c in header_cells)

    if is_own_template:
        return _import_own_template(ws)

    header_info = find_statement_header(ws)
    if not header_info:
        return jsonify({
            "error": "엑셀에서 날짜·금액 열을 찾지 못했습니다. 카드사에서 받은 파일 원본 그대로 올려주세요 "
                     "(헤더에 '이용일자'나 '승인금액' 같은 표현이 있어야 자동으로 인식됩니다)."
        }), 400

    return _import_statement(ws, header_info)


init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    app.run(host="0.0.0.0", port=port, debug=False)
